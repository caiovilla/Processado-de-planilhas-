from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import tkinter as tk
from PIL import Image
from tkinter import filedialog, messagebox
import os

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    return texto.replace("\n", "").strip().lower()

def buscar_valor_por_chave(ws, chave):
    chave_normalizada = normalizar_texto(chave)
    for row in ws.iter_rows():
        for cell in row:
            if normalizar_texto(cell.value) == chave_normalizada:
                return cell.offset(column=1).value
    return None

class ModelocProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador de Cadastro modelo c")

        self.dados_file = None
        self.modelo_file = None
        self.sheet_dados_name = tk.StringVar()
        self.sheet_modelo_name = tk.StringVar()
        self.sheet_valor1_name = tk.StringVar()
        self.sheet_valor2_name = tk.StringVar()
        

        self.create_widgets()
    
    def create_widgets(self):
        tk.Label(self.root, text="Ficha de cadastro (Excel):").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.dados_label = tk.Label(self.root, text="Nenhum arquivo selecionado", width=50)
        self.dados_label.grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.root, text="Selecionar", command=self.select_dados).grid(row=0, column=2, padx=5, pady=5)
        
        
        tk.Label(self.root, text="Aba de dados:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.dropdown_dados = tk.OptionMenu(self.root, self.sheet_dados_name, "")
        self.dropdown_dados.grid(row=1, column=1, padx=5, pady=5)
        
        
        tk.Label(self.root, text="Arquivo de modelo (Excel):").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.modelo_label = tk.Label(self.root, text="Nenhum arquivo selecionado", width=50)
        self.modelo_label.grid(row=2, column=1, padx=5, pady=5)
        tk.Button(self.root, text="Selecionar", command=self.select_modelo).grid(row=2, column=2, padx=5, pady=5)
        
        
        tk.Label(self.root, text="Aba do modelo:").grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.dropdown_modelo = tk.OptionMenu(self.root, self.sheet_modelo_name, "")
        self.dropdown_modelo.grid(row=3, column=1, padx=5, pady=5)
        
        
        tk.Button(self.root, text="Processar", command=self.processar).grid(row=7, column=0, columnspan=3, pady=10)
        
    def select_dados(self):
        self.dados_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not self.dados_file:
            messagebox.showwarning("Aviso", "Nenhum arquivo de dados selecionado.")
            return

        try:
            wb = load_workbook(self.dados_file, read_only=True)
            sheets = wb.sheetnames
            if not sheets:
                raise ValueError("O arquivo de dados não possui abas.")
        
            self.dados_label.config(text=os.path.basename(self.dados_file))

       
            m_dados = self.dropdown_dados["menu"]
            m_dados.delete(0, "end")
            for sheet in sheets:
                m_dados.add_command(label=sheet, command=lambda value=sheet: self.sheet_dados_name.set(value))
            self.sheet_dados_name.set(sheets[0])

        
            for dropdown, var in [(self.dropdown_valor1, self.sheet_valor1_name), (self.dropdown_valor2, self.sheet_valor2_name)]:
                m = dropdown["menu"]
                m.delete(0, "end")
                for sheet in sheets:
                    m.add_command(label=sheet, command=lambda value=sheet, v=var: v.set(value))
                var.set(sheets[0])

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar arquivo de dados:\n{str(e)}")
    
    def select_modelo(self):
        self.modelo_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.modelo_file:
            self.modelo_label.config(text=os.path.basename(self.modelo_file))
            wb = load_workbook(self.modelo_file, read_only=True)
            sheets = wb.sheetnames
            m = self.dropdown_modelo["menu"]
            m.delete(0, "end")
            for sheet in sheets:
                m.add_command(label=sheet, command=lambda value=sheet: self.sheet_modelo_name.set(value))
            if sheets:
                self.sheet_modelo_name.set(sheets[0])
    
    def processar(self):
        if not self.dados_file or not os.path.exists(self.dados_file):
            messagebox.showerror("Erro", "Arquivo de dados não selecionado ou inválido.")
            return
        if not self.modelo_file or not os.path.exists(self.modelo_file):
            messagebox.showerror("Erro", "Arquivo de modelo não selecionado ou inválido.")
            return

        if not self.dados_file or not self.modelo_file:
            messagebox.showerror("Erro", "Por favor, selecione os arquivos necessários.")
            return
        
        wb_dados = load_workbook(self.dados_file, data_only=True)
        ws_dados = wb_dados[self.sheet_dados_name.get()]
        wb_modelo = load_workbook(self.modelo_file)
        ws_modelo = wb_modelo[self.sheet_modelo_name.get()]
        
        dados_extraidos = {
            "Código de Barras/EAN": buscar_valor_por_chave(ws_dados, "Código de Barras/EAN"),
            "Nome do Produto": buscar_valor_por_chave(ws_dados, "Nome do Produto"),
            "Tecnologia ou ingredientes chaves (Princípio Ativo)": buscar_valor_por_chave(ws_dados, "Tecnologia ou ingredientes chaves (Princípio Ativo)"),
            "Marca": buscar_valor_por_chave(ws_dados, "Marca"),
            "Tipo de Embalagem": buscar_valor_por_chave(ws_dados, "Tipo de Embalagem"),
            "Indicação (Necessidade/Tipo)": buscar_valor_por_chave(ws_dados, "Indicação (Necessidade/Tipo)"),
        }

        ws_modelo["H2"] = dados_extraidos["Código de Barras/EAN"]
        ws_modelo["A2"] = dados_extraidos["Nome do Produto"]
        ws_modelo["B2"] = dados_extraidos["Marca"]
        ws_modelo["D2"] = dados_extraidos["Tecnologia ou ingredientes chaves (Princípio Ativo)"]
        ws_modelo["E2"] = dados_extraidos["Tipo de Embalagem"]
        
    
        
        default_filename = f"Cadastro_modelc_.xlsx"
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename,
            title="Salvar planilha como..."
        )

        if save_path:
            wb_modelo.save(save_path)
            messagebox.showinfo("Sucesso", f"Planilha gerada com sucesso:\n{save_path}")
        else:
            messagebox.showinfo("Cancelado", "Salvamento cancelado pelo usuário.")

def iniciar_app():
    root = tk.Tk()
    app = ModelocProcessorApp(root)
    root.mainloop()
                
    
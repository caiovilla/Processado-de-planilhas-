from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import datetime
from dateutil.parser import parse

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    return texto.replace("\n", "").strip().lower()

def buscar_valor_por_chave(ws, chave):
    chave_normalizada = normalizar_texto(chave)
    for row in ws.iter_rows():
        for cell in row:
            if normalizar_texto(cell.value) == chave_normalizada:
                return cell.offset(column=1).value
    return None

def buscar_valor_chave_especial(ws, chave):
    chave_especial_normalizada = normalizar_texto(chave)
    
    for row in ws.iter_rows():
        for cell in row:
            valor_celula = cell.value
            if valor_celula is not None and normalizar_texto(valor_celula) == chave_especial_normalizada:
               
                col_atual = cell.column  
                linha_atual = cell.row
                
                for offset_col in range(1, 10):  
                    valor_frente = ws.cell(row=linha_atual, column=col_atual + offset_col).value
                    if valor_frente is not None and str(valor_frente).strip() != "":
                        return valor_frente
               
                return None
    
    return None 


def buscar_valor_nome_anvisa(ws):
    texto_referencia = "nome anvisa"
    for row in ws.iter_rows():
        for cell in row:
            if normalizar_texto(cell.value) == texto_referencia:
                cell_direita = cell.offset(column=1)
                cell_alvo = ws.cell(row=cell_direita.row - 1, column=cell_direita.column)
                return cell_alvo.value
    return None

class ModeloaExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador de Cadastro a")

        self.dados_file = None
        self.modelo_file = None
        self.sheet_dados_name = tk.StringVar()
        self.sheet_modelo_name = tk.StringVar()
        self.sheet_valor1_name = tk.StringVar()
        self.sheet_valor2_name = tk.StringVar()
        

        self.create_widgets()

    def create_widgets(self):
    
        tk.Label(self.root, text="Ficha de cadastro (Excel):").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.dados_label = tk.Label(self.root, text="Nenhum arquivo selecionado", width=50)
        self.dados_label.grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.root, text="Selecionar", command=self.select_dados).grid(row=0, column=2, padx=5, pady=5)

   
        tk.Label(self.root, text="Aba de dados:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.dropdown_dados = tk.OptionMenu(self.root, self.sheet_dados_name, "")
        self.dropdown_dados.grid(row=1, column=1, padx=5, pady=5)

    
        tk.Label(self.root, text="Arquivo de modelo (Excel):").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.modelo_label = tk.Label(self.root, text="Nenhum arquivo selecionado", width=50)
        self.modelo_label.grid(row=2, column=1, padx=5, pady=5)
        tk.Button(self.root, text="Selecionar", command=self.select_modelo).grid(row=2, column=2, padx=5, pady=5)

   
        tk.Label(self.root, text="Aba do modelo:").grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.dropdown_modelo = tk.OptionMenu(self.root, self.sheet_modelo_name, "")
        self.dropdown_modelo.grid(row=3, column=1, padx=5, pady=5)

   
        tk.Label(self.root, text="Imagem (JPG,JPEG,PNG):").grid(row=4, column=0, sticky='w', padx=5, pady=5)
        self.imagem_label = tk.Label(self.root, text="Nenhum arquivo selecionado", width=50)
        self.imagem_label.grid(row=4, column=1, padx=5, pady=5)
        tk.Button(self.root, text="Selecionar", command=self.select_imagem).grid(row=4, column=2, padx=5, pady=5)

    
        tk.Label(self.root, text="Aba para Valor Consumidor:").grid(row=5, column=0, sticky='w', padx=5, pady=5)
        self.dropdown_valor1 = tk.OptionMenu(self.root, self.sheet_valor1_name, "")
        self.dropdown_valor1.grid(row=5, column=1, padx=5, pady=5)

        tk.Label(self.root, text="Aba para Valor Lista:").grid(row=6, column=0, sticky='w', padx=5, pady=5)
        self.dropdown_valor2 = tk.OptionMenu(self.root, self.sheet_valor2_name, "")
        self.dropdown_valor2.grid(row=6, column=1, padx=5, pady=5)

    
        tk.Button(self.root, text="Processar", command=self.processar).grid(row=7, column=0, columnspan=3, pady=10)



    def select_dados(self):
        self.dados_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not self.dados_file:
            messagebox.showwarning("Aviso", "Nenhum arquivo de dados selecionado.")
            return

        try:
            wb = load_workbook(self.dados_file, read_only=True)
            sheets = wb.sheetnames
            if not sheets:
                raise ValueError("O arquivo de dados não possui abas.")
        
            self.dados_label.config(text=os.path.basename(self.dados_file))

       
            m_dados = self.dropdown_dados["menu"]
            m_dados.delete(0, "end")
            for sheet in sheets:
                m_dados.add_command(label=sheet, command=lambda value=sheet: self.sheet_dados_name.set(value))
            self.sheet_dados_name.set(sheets[0])

        
            for dropdown, var in [(self.dropdown_valor1, self.sheet_valor1_name), (self.dropdown_valor2, self.sheet_valor2_name)]:
                m = dropdown["menu"]
                m.delete(0, "end")
                for sheet in sheets:
                    m.add_command(label=sheet, command=lambda value=sheet, v=var: v.set(value))
                var.set(sheets[0])

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar arquivo de dados:\n{str(e)}")


      
    def select_modelo(self):
        self.modelo_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.modelo_file:
            self.modelo_label.config(text=os.path.basename(self.modelo_file))
            wb = load_workbook(self.modelo_file, read_only=True)
            sheets = wb.sheetnames
            m = self.dropdown_modelo["menu"]
            m.delete(0, "end")
            for sheet in sheets:
                m.add_command(label=sheet, command=lambda value=sheet: self.sheet_modelo_name.set(value))
            if sheets:
                self.sheet_modelo_name.set(sheets[0])
    
    def select_imagem(self):
        path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.png;*.jpeg")])
        if not path:
            messagebox.showwarning("Aviso", "Nenhuma imagem selecionada.")
            self.imagem_file = None
            self.imagem_label.config(text="Nenhum arquivo selecionado")
            return
        try:
            img = Image.open(path)
            img.verify()
            img = Image.open(path)
            
            if img.mode in("RGBA", "P"):
                img = img.convert("RGB")
            
            img.thumbnail((500, 500))
            
            resized_path = "temp_resized.jpg"
            img.save(resized_path)
            
            self.imagem_file = resized_path
            self.imagem_label.config(text=os.path.basename(path))
            
        except Exception as e:
             self.imagem_file = None
             self.imagem_label.config(text="Imagem inválida")
             messagebox.showerror("Erro ao carregar imagem", f"A imagem selecionada é inválida ou está corrompida:\n{str(e)}")


    def processar(self):
        if not self.dados_file or not os.path.exists(self.dados_file):
            messagebox.showerror("Erro", "Arquivo de dados não selecionado ou inválido.")
            return
        if not self.modelo_file or not os.path.exists(self.modelo_file):
            messagebox.showerror("Erro", "Arquivo de modelo não selecionado ou inválido.")
            return
        if not self.imagem_file or not os.path.exists(self.imagem_file):
            messagebox.showerror("Erro", "Imagem não selecionada ou inválida.")
            return

        if not self.sheet_dados_name.get() or not self.sheet_modelo_name.get():
            messagebox.showerror("Erro", "Selecione as abas corretamente.")
            return

        if not self.dados_file or not self.modelo_file:
            messagebox.showerror("Erro", "Por favor, selecione os arquivos necessários.")
            return

        wb_dados = load_workbook(self.dados_file, data_only=True)
        ws_dados = wb_dados[self.sheet_dados_name.get()]
        wb_modelo = load_workbook(self.modelo_file)
        ws_modelo = wb_modelo[self.sheet_modelo_name.get()]
        ws_valor1 = wb_dados[self.sheet_valor1_name.get()]
        ws_valor2 = wb_dados[self.sheet_valor2_name.get()]
        
        

        dados_extraidos = {
            "Código de Barras/EAN": buscar_valor_por_chave(ws_dados, "Código de Barras/EAN"),
            "Nome Completo do Produto": buscar_valor_por_chave(ws_dados, "Nome Completo do Produto"),
            "Marca": buscar_valor_por_chave(ws_dados, "Marca"),
            "Tecnologia ou ingredientes chaves (Princípio Ativo)": buscar_valor_por_chave(ws_dados, "Tecnologia ou ingredientes chaves (Princípio Ativo)"),
            "SÃO PAULO": buscar_valor_por_chave(ws_valor2, "SÃO PAULO"),
            "Preço consumidor": buscar_valor_chave_especial(ws_valor1,"Preço consumidor"),
            "Tipo de Embalagem": buscar_valor_por_chave(ws_dados, "Tipo de Embalagem"),
            "Data de lançamento (Sell In)": buscar_valor_por_chave(ws_dados, "Data de lançamento (Sell In)"),
            "Indicação (Necessidade/Tipo)": buscar_valor_por_chave(ws_dados, "Indicação (Necessidade/Tipo)"),
            "Kit": buscar_valor_por_chave(ws_dados, "Kit"),
            "Detalhamento do KIT": buscar_valor_por_chave(ws_dados,"Detalhamento do KIT")
        }

        ws_modelo["G11"] = dados_extraidos["Código de Barras/EAN"]
        ws_modelo["G13"] = dados_extraidos["Nome Completo do Produto"]
        ws_modelo["G17"] = dados_extraidos["Nome Completo do Produto"]
        ws_modelo["G15"] = dados_extraidos["Marca"]
        ws_modelo["G21"] = dados_extraidos["Marca"]
        ws_modelo["G19"] = dados_extraidos["Marca"]
        ws_modelo["G23"] = dados_extraidos["Tecnologia ou ingredientes chaves (Princípio Ativo)"]
        ws_modelo["G27"] = dados_extraidos["Preço consumidor"]
        ws_modelo["G25"] = dados_extraidos["SÃO PAULO"]
        ws_modelo["G29"] = dados_extraidos["Tipo de Embalagem"]
        ws_modelo["G31"] = dados_extraidos["Data de lançamento (Sell In)"]
        ws_modelo["G33"] = dados_extraidos["Indicação (Necessidade/Tipo)"]
        ws_modelo["G37"] = dados_extraidos["Kit"]
        ws_modelo["G39"]= dados_extraidos["Detalhamento do KIT"]
        
        img = ExcelImage(self.imagem_file)
        ws_modelo.add_image(img, "C9")

        nome = dados_extraidos["Nome Completo do Produto"] or "sem_nome"
        nome = "".join(c for c in nome if c.isalnum() or c in " _-").strip()  

        default_filename = f"Cadastro_modeloa_{nome}.xlsx"


        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename,
            title="Salvar planilha como..."
        )

        if save_path:
            wb_modelo.save(save_path)
            messagebox.showinfo("Sucesso", f"Planilha gerada com sucesso:\n{save_path}")
        else:
            messagebox.showinfo("Cancelado", "Salvamento cancelado pelo usuário.")

def iniciar_app():
    root = tk.Tk()
    app = ModeloaExcelProcessorApp(root)
    root.mainloop()
